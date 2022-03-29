# from django.apps import AppConfig
import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "server.settings")
import django
django.setup()
from django.core.wsgi import get_wsgi_application
from openpyxl import Workbook, load_workbook
from menus.serializers import FoodSerializer

import re

# FILE_NAME = "datas/통합 식품영양성분DB.xlsx"
# SHEET_NAME = "원본"
FILE_NAME = "datas/통합 식품영양성분DB_음식_20220308.xlsx"
SHEET_NAME = "Data"

# 엑셀 파일의 필요 속성 13개
xls_properties = ["NO", "상용제품", "식품명", "식품대분류", "식품상세분류", "1회제공량", "에너지(㎉)", "총당류(g)", "탄수화물(g)", "단백질(g)"
, "지방(g)", "콜레스테롤(㎎)", "총 포화 지방산(g)", "고기", "야채", "해산물", "맵기", "기름기"]

# Food 테이블의 13개 속성 중 image를 제외한 12가지 매칭 + 품목대표인지 확인을 위해 "상용제품" 속성 추가 필요 -> "상용제품" : commercialFood
table_properties = {"NO" : "id", "상용제품" : "commercialFood", "식품명" : "foodName", "식품대분류" : "foodCategory", "식품상세분류" : "foodDetailCategory"
, "1회제공량" : "servingSize", "에너지(㎉)" : "foodKcal", "총당류(g)" : "sugar", "탄수화물(g)" : "carbohydrate"
, "단백질(g)" : "protein", "지방(g)" : "fat", "콜레스테롤(㎎)" : "cholesterol", "총 포화 지방산(g)" : "fattyAcid"
, "고기" : "meat", "야채" : "vegetable", "해산물" : "seafood", "맵기" : "spicy", "기름기" : "oily"}

class XlsxParser():
    def __init__(self):
        self.load_wb = load_workbook(filename= FILE_NAME, data_only=True)
        self.load_ws = self.load_wb[SHEET_NAME]
        self.match_Properties_To_Index()
        print("data amount :", (self.load_ws.max_row-4))

    def match_Properties_To_Index(self):
        # 엑셀 속성의 컬럼 index와 속성명을 매칭
        self.xls_index_to_values = {}
        for tuples in self.load_ws['4']:
            # print(tuples.row, tuples.column, tuples.value)
            if tuples.value in xls_properties:
                self.xls_index_to_values[tuples.column] = tuples.value
        # print("data amount :", len(self.load_ws.rows))
    
    def input_datas(self):
        # all_foods = []
        cnt = 0;
        fail_cnt = 0;
        p = re.compile("[0-9].[0-9]")
        for row in self.load_ws.rows:
            if row[0].row < 5: continue;
            food = {}
            for cell in row:
                if cell.column in self.xls_index_to_values.keys():
                    
                    try:
                        if cell.value == "-":
                            f_value = 0
                        elif "미만" in cell.value:
                            f_value = 0
                        # elif p.match(cell.value) is not None:
                        #     f_value = round(float(cell.value), 2)
                        elif type(cell.value) is not int:
                            f_value = round(float(cell.value), 2)
                        else:
                            f_value = cell.value.strip()
                        food[table_properties[self.xls_index_to_values[cell.column]]] = f_value
                    except:
                        food[table_properties[self.xls_index_to_values[cell.column]]] = cell.value
            
            try:
                serializer = FoodSerializer(data=food)
                if serializer.is_valid(raise_exception=True):
                    serializer.save()
                cnt += 1
            except:
                print(food)
                fail_cnt += 1
            
            if cnt%1000 == 0:
                print(cnt, "datas accessed. . .")
            # all_foods.append(food)
        # print(all_foods)
        return cnt, fail_cnt

def main():
    # p = re.compile("[0-9].[0-9]")
    # print(p.match("1.9"))
    # print(p.match("a.b"))
    # print(p.match("1.99"))
    print("start db_uploader")

    print("initializing db_uploader")
    xlsxParser = XlsxParser()
    print("initializing finished")

    xlsxParser.match_Properties_To_Index()

    print("start input_datas")
    cnt, fail_cnt = xlsxParser.input_datas()
    print("input_datas finished")

    print("-----------------result-----------------")
    print("Successed :", (cnt), ", Failed :", fail_cnt, ", Total :", (cnt+fail_cnt))
    print("----------------------------------------")

    print("db_uploader finished")


if __name__ == '__main__':
    main()
